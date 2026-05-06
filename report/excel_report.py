"""
Генератор Excel-отчёта с несколькими листами.
"""
from __future__ import annotations

import os
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from config import AppConfig

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="0078D4")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ALT_FILL = PatternFill(fill_type="solid", fgColor="EBF3FB")
_BORDER_SIDE = Side(style="thin", color="CCCCCC")
_CELL_BORDER = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE,
    top=_BORDER_SIDE, bottom=_BORDER_SIDE,
)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def _style_sheet(ws, df: pd.DataFrame) -> None:
    """Применяет стили к листу с данными DataFrame."""
    # Заголовки
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _ALIGN_CENTER
        cell.border = _CELL_BORDER

    # Данные
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = _CELL_BORDER
            if fill:
                cell.fill = fill
            # Выравнивание: первая колонка — влево, числа — вправо
            if col_idx == 1:
                cell.alignment = _ALIGN_LEFT
            else:
                cell.alignment = _ALIGN_RIGHT

    # Авто-ширина колонок
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col_name)),
            *(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # Закрепить первую строку
    ws.freeze_panes = "A2"


class ExcelReport:
    """Генерирует Excel-отчёт с несколькими листами."""

    def __init__(
        self,
        summary_df: pd.DataFrame,
        timeline_df: pd.DataFrame,
        config: AppConfig,
    ) -> None:
        self.summary_df = summary_df
        self.timeline_df = timeline_df
        self.config = config

    def generate(self, output_path: str) -> None:
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

        col_labels_summary = {
            "person": "Участник",
            "commits": "Коммиты",
            "lines_added": "+ строк",
            "lines_deleted": "− строк",
            "prs_created": "PR создано",
            "prs_merged": "PR влито",
            "prs_abandoned": "PR закрыто",
            "reviews_given": "Ревью",
            "reviews_approved": "Одобрено",
            "reviews_suggestions": "С замеч.",
            "reviews_rejected": "Откл.",
            "pr_comments": "Комментарии",
            "wi_created": "WI создано",
            "wi_closed": "WI закрыто",
            "wi_assigned": "WI назначено",
        }

        col_labels_timeline = {
            "person": "Участник",
            "week": "Неделя",
            "commits": "Коммиты",
            "prs_created": "PR создано",
            "prs_merged": "PR влито",
            "reviews_given": "Ревью",
            "pr_comments": "Комментарии",
            "wi_created": "WI создано",
            "wi_closed": "WI закрыто",
        }

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Лист "Сводка"
            summary_out = self.summary_df.rename(columns=col_labels_summary)
            summary_out.to_excel(writer, sheet_name="Сводка", index=False)

            # Лист "По неделям"
            timeline_out = self.timeline_df.rename(columns=col_labels_timeline)
            timeline_out.to_excel(writer, sheet_name="По неделям", index=False)

            # Листы по участникам (если < 20)
            persons = list(self.summary_df["person"]) if not self.summary_df.empty else []
            if len(persons) < 20 and not self.timeline_df.empty:
                for person in persons:
                    sheet_name = self._safe_sheet_name(person)
                    person_df = self.timeline_df[self.timeline_df["person"] == person].copy()
                    person_df = person_df.drop(columns=["person"], errors="ignore")
                    person_df = person_df.rename(columns=col_labels_timeline)
                    person_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Применяем стили через openpyxl (pandas ExcelWriter оставляет стандартный вид)
        wb = load_workbook(output_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 2:
                continue
            # Перестраиваем стили
            self._apply_styles(ws)

        # Добавляем информационный лист
        self._add_info_sheet(wb)

        wb.save(output_path)

    # ------------------------------------------------------------------

    def _apply_styles(self, ws) -> None:
        """Применяет стили к уже заполненному листу."""
        # Заголовки (строка 1)
        for cell in ws[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _ALIGN_CENTER
            cell.border = _CELL_BORDER

        # Данные
        for row_idx in range(2, ws.max_row + 1):
            fill = _ALT_FILL if row_idx % 2 == 0 else None
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = _CELL_BORDER
                if fill:
                    cell.fill = fill
                cell.alignment = _ALIGN_LEFT if col_idx == 1 else _ALIGN_RIGHT

        # Авто-ширина
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

        ws.freeze_panes = "A2"

    def _add_info_sheet(self, wb) -> None:
        """Добавляет лист с метаданными отчёта."""
        ws = wb.create_sheet("Информация", 0)
        info = [
            ("Параметр", "Значение"),
            ("Период с", self.config.date_from.strftime("%d.%m.%Y")),
            ("Период по", self.config.date_to.strftime("%d.%m.%Y")),
            ("Проекты", ", ".join(self.config.projects)),
            ("Участников", str(len(self.summary_df) if not self.summary_df.empty else 0)),
        ]
        for row_idx, (key, value) in enumerate(info, start=1):
            ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 40

    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        """Обрезает и очищает имя листа для Excel (макс. 31 символ)."""
        invalid = r'\/*?:[]]'
        for ch in invalid:
            name = name.replace(ch, "_")
        return name[:31]
